# envios/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q, Sum
from django.views.decorators.http import require_GET
from django.db import transaction
from datetime import datetime
import json
from .models import Producto, Stock, Bodega, GuiaEnvio, GuiaEnvioProducto, Campaign
from .forms import ProductoForm, StockForm, BodegaForm, GuiaEnvioForm
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal
from django.http import JsonResponse, HttpResponse
import csv
import traceback

# --- Imports for Export Functionality ---
from django.utils import timezone
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from io import BytesIO
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.urls import reverse

# --- Vistas de Producto ---
def listar_productos(request):
    search_query = request.GET.get('search', '').strip()
    productos = Producto.objects.all()
    if search_query:
        try:
            id_query = int(search_query)
            productos = productos.filter(Q(id=id_query))
        except ValueError:
            productos = productos.filter(
                Q(nombre__icontains=search_query) | Q(referencia__icontains=search_query)
            )
    return render(request, 'productos/productos.html', {'productos': productos, 'search_query': search_query})

def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto creado correctamente.')
            return redirect('listar_productos')
    else:
        form = ProductoForm()
    return render(request, 'productos/crear_producto.html', {'form': form})

def ver_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    return render(request, 'productos/ver_producto.html', {'producto': producto})

def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('listar_productos')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'productos/editar_producto.html', {'form': form})

def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    if request.method == 'POST':
        producto.delete()
        messages.success(request, f'Producto "{producto.nombre}" eliminado correctamente.')
        return redirect('listar_productos')
    return render(request, 'productos/eliminar_producto.html', {'producto': producto})

# --- Vistas de Stock ---
def listar_stocks(request):
    stocks_qs = Stock.objects.select_related('producto', 'bodega').all().order_by('producto__nombre', 'bodega__nombre')

    # Obtener parámetros de filtro del GET request
    search_query = request.GET.get('search', '').strip()
    estado_filter = request.GET.get('estado', '')
    bodega_id_filter = request.GET.get('bodega_id', '')
    producto_id_filter = request.GET.get('producto_id', '') # NUEVO FILTRO
    export_format = request.GET.get('export_format', '')

    # Aplicar filtros al queryset
    if search_query:
        q_objects = Q(producto__nombre__icontains=search_query) | \
                    Q(producto__referencia__icontains=search_query)
        try:
            search_id = int(search_query)
            q_objects |= Q(id=search_id) # Buscar por ID de Stock
        except ValueError:
            pass
        stocks_qs = stocks_qs.filter(q_objects)

    if estado_filter:
        stocks_qs = stocks_qs.filter(estado=estado_filter)

    if bodega_id_filter:
        try:
            bodega_id_val = int(bodega_id_filter)
            stocks_qs = stocks_qs.filter(bodega__id=bodega_id_val)
        except ValueError:
            if bodega_id_filter: # Solo mostrar advertencia si se intentó filtrar con valor no numérico
                messages.warning(request, "ID de bodega inválido.")
    
    if producto_id_filter: # NUEVO FILTRO
        try:
            producto_id_val = int(producto_id_filter)
            stocks_qs = stocks_qs.filter(producto__id=producto_id_val)
        except ValueError:
            if producto_id_filter: # Solo mostrar advertencia si se intentó filtrar
                 messages.warning(request, "ID de producto inválido.")


    # ---- Lógica de Exportación (sin cambios aquí, ya usa stocks_qs filtrado) ----
    if export_format:
        filename_timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        
        if export_format == 'txt':
            response = HttpResponse(content_type='text/plain; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="stocks_export_{filename_timestamp}.txt"'
            writer = csv.writer(response, delimiter='\t')
            writer.writerow(['ID Stock', 'Producto', 'Referencia Prod.', 'Bodega', 'Stock Actual', 'Umbral Mínimo', 'Estado'])
            for stock_item in stocks_qs:
                writer.writerow([
                    stock_item.id,
                    stock_item.producto.nombre,
                    stock_item.producto.referencia or '',
                    stock_item.bodega.nombre,
                    stock_item.stock_actual,
                    stock_item.umbral_minimo,
                    stock_item.get_estado_display()
                ])
            return response

        elif export_format == 'excel':
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="stocks_export_{filename_timestamp}.xlsx"'
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Stocks"
            headers = ['ID Stock', 'Producto', 'Referencia Prod.', 'Bodega', 'Stock Actual', 'Umbral Mínimo', 'Estado']
            for col_num, header_title in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num, value=header_title)
                cell.font = Font(bold=True)
                sheet.column_dimensions[get_column_letter(col_num)].width = 25
            for row_num, stock_item in enumerate(stocks_qs, 2):
                sheet.cell(row=row_num, column=1, value=stock_item.id)
                sheet.cell(row=row_num, column=2, value=stock_item.producto.nombre)
                sheet.cell(row=row_num, column=3, value=stock_item.producto.referencia or '')
                sheet.cell(row=row_num, column=4, value=stock_item.bodega.nombre)
                sheet.cell(row=row_num, column=5, value=stock_item.stock_actual)
                sheet.cell(row=row_num, column=6, value=stock_item.umbral_minimo)
                sheet.cell(row=row_num, column=7, value=stock_item.get_estado_display())
            workbook.save(response)
            return response
        # PDF omitido por brevedad, añadir si es necesario

    # ---- Renderizado HTML Normal ----
    if not stocks_qs.exists() and (search_query or estado_filter or bodega_id_filter or producto_id_filter): # AÑADIDO producto_id_filter
        messages.info(request, "No hay registros de stock que coincidan con los filtros aplicados.")
    elif not stocks_qs.exists():
         messages.info(request, "No hay registros de stock disponibles.")

    context = {
        'stocks': stocks_qs,
        'bodegas_all': Bodega.objects.all().order_by('nombre'),
        'productos_all': Producto.objects.all().order_by('nombre'), # NUEVO: para el desplegable de productos
        'search_query': search_query,
        'estado_selected': estado_filter,
        'bodega_id_selected': bodega_id_filter,
        'producto_id_selected': producto_id_filter, # NUEVO: para preseleccionar producto
    }
    return render(request, 'stocks/stocks.html', context)

def crear_stock(request):
    if request.method == 'POST':
        form = StockForm(request.POST)
        if form.is_valid():
            try:
                nuevo_stock = form.save()
                messages.success(request, f"Stock para {nuevo_stock.producto.nombre} creado!")
                return redirect('listar_stocks')
            except Exception as e:
                messages.error(request, f"Error al guardar: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
    else:
        form = StockForm()
    return render(request, 'stocks/crear.html', {'form': form, 'productos': Producto.objects.all(), 'bodegas': Bodega.objects.all()})

def editar_stock(request, id):
    stock = get_object_or_404(Stock, id=id)
    if request.method == 'POST':
        form = StockForm(request.POST, instance=stock)
        if form.is_valid():
            form.save()
            messages.success(request, 'Stock actualizado.')
            return redirect('listar_stocks')
    else:
        form = StockForm(instance=stock)
    return render(request, 'stocks/editar_stock.html', {'form': form})

def eliminar_stock(request, id):
    stock = get_object_or_404(Stock, id=id)
    if request.method == 'POST':
        stock.delete()
        messages.success(request, 'Stock eliminado.')
        return redirect('listar_stocks')
    return render(request, 'stocks/eliminar_stock.html', {'stock': stock})

def ver_stock(request, id):
    stock = get_object_or_404(Stock, id=id)
    return render(request, 'stocks/ver_stock.html', {'stock': stock})

# --- Vistas de Bodega ---
# envios/views.py

def listar_bodegas(request):
    search_query = request.GET.get('search', '').strip()
    bodegas = Bodega.objects.all().order_by('nombre') # Añadido order_by

    if search_query:
        try:
            id_query = int(search_query)
            bodegas = bodegas.filter(Q(id=id_query))
        except ValueError:
            bodegas = bodegas.filter(
                Q(nombre__icontains=search_query) |
                Q(direccion__icontains=search_query) |
                Q(telefono__icontains=search_query) |
                Q(ciudad__icontains=search_query)  # AÑADIDO FILTRO POR CIUDAD
            )
    
    context = {
        'bodegas': bodegas, 
        'search_query': search_query
    }
    return render(request, 'bodegas/listar_bodegas.html', context)

def crear_bodega(request):
    if request.method == 'POST':
        form = BodegaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Bodega creada.')
            return redirect('listar_bodegas')
    else:
        form = BodegaForm()
    return render(request, 'bodegas/crear_bodega.html', {'form': form})

def editar_bodega(request, id):
    bodega = get_object_or_404(Bodega, id=id)
    if request.method == 'POST':
        form = BodegaForm(request.POST, instance=bodega)
        if form.is_valid():
            form.save()
            return redirect('listar_bodegas')
    else:
        form = BodegaForm(instance=bodega)
    return render(request, 'bodegas/editar_bodega.html', {'form': form})

def eliminar_bodega(request, id):
    bodega = get_object_or_404(Bodega, id=id)
    if request.method == 'POST':
        bodega.delete()
        return redirect('listar_bodegas')
    return render(request, 'bodegas/eliminar_bodega.html', {'bodega': bodega})

def ver_bodega(request, id):
    bodega = get_object_or_404(Bodega, id=id)
    return render(request, 'bodegas/ver_bodega.html', {'bodega': bodega})


# --- API VIEWS ---
@require_GET
def search_productos(request):
    search_query = request.GET.get('search', '').strip()
    productos_qs = Producto.objects.all()

    if search_query:
        productos_qs = productos_qs.filter(
            Q(nombre__icontains=search_query) | Q(referencia__icontains=search_query)
        )[:10]
    else:
        productos_qs = Producto.objects.none()

    results = []
    for p in productos_qs:
        results.append({
            'id': p.id,
            'nombre': p.nombre,
            'referencia': p.referencia if p.referencia else '',
            'precio': str(p.precio),
            'imagen': p.imagen.url if p.imagen else None
        })
    return JsonResponse(results, safe=False)

@require_GET
def producto_detail_api(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    data = {
        'id': producto.id,
        'nombre': producto.nombre,
        'referencia': producto.referencia,
        'precio': str(producto.precio),
        'imagen': producto.imagen.url if producto.imagen else None,
    }
    return JsonResponse(data)

# --- GUIA DE ENVIO VIEWS ---
@transaction.atomic
def crear_guia(request):
    tabla_productos_data_json = '[]'
    if request.method == 'POST':
        form = GuiaEnvioForm(request.POST)
        productos_en_post_para_repopular = []
        i = 0
        while f'productos[{i}][id]' in request.POST:
            try:
                prod_id = request.POST[f'productos[{i}][id]']
                prod_obj = Producto.objects.get(id=prod_id)
                productos_en_post_para_repopular.append({
                    'id': prod_id,
                    'nombre': request.POST.get(f'productos[{i}][nombre]', prod_obj.nombre),
                    'precio': request.POST.get(f'productos[{i}][precio]', str(prod_obj.precio)),
                    'cantidad': request.POST[f'productos[{i}][cantidad]'],
                    'imagen': prod_obj.imagen.url if prod_obj.imagen else 'https://via.placeholder.com/50'
                })
            except Exception:
                pass 
            i += 1
        if productos_en_post_para_repopular:
            tabla_productos_data_json = json.dumps(productos_en_post_para_repopular)

        if form.is_valid():
            try:
                guia = form.save(commit=False)
                items_para_crear_en_db = []
                idx = 0
                hay_productos_en_tabla = False
                primer_producto_obj_de_tabla = None
                primera_cantidad_de_tabla = 1

                while f'productos[{idx}][id]' in request.POST:
                    hay_productos_en_tabla = True
                    producto_id_str = request.POST[f'productos[{idx}][id]']
                    cantidad_str = request.POST[f'productos[{idx}][cantidad]']
                    precio_unitario_str = request.POST.get(f'productos[{idx}][precio]', '0')

                    try:
                        producto_id = int(producto_id_str)
                        cantidad_item = int(cantidad_str)
                        precio_unitario = Decimal(precio_unitario_str)
                        if cantidad_item <= 0: raise ValueError("La cantidad debe ser positiva.")
                        if precio_unitario < 0: raise ValueError("El precio no puede ser negativo.")
                    except ValueError as ve:
                        messages.error(request, f"Error en datos de producto {idx+1} de la tabla: {ve}")
                        return render(request, 'guias/crear_guia.html', {'form': form, 'error_productos': True, 'tabla_productos_data_json': tabla_productos_data_json})

                    producto_obj_item = get_object_or_404(Producto, id=producto_id)
                    if idx == 0:
                        primer_producto_obj_de_tabla = producto_obj_item
                        primera_cantidad_de_tabla = cantidad_item

                    items_para_crear_en_db.append(GuiaEnvioProducto(
                        producto=producto_obj_item,
                        cantidad=cantidad_item,
                        precio_unitario_en_guia=precio_unitario
                    ))
                    idx += 1
                
                if not hay_productos_en_tabla:
                    messages.error(request, "Debe agregar al menos un producto a la tabla.")
                    return render(request, 'guias/crear_guia.html', {'form': form, 'error_productos': True, 'tabla_productos_data_json': tabla_productos_data_json})

                if primer_producto_obj_de_tabla:
                    guia.producto = primer_producto_obj_de_tabla
                    guia.cantidad = primera_cantidad_de_tabla
                else: 
                    messages.error(request, "Error: No se pudo determinar el producto principal.")
                    return render(request, 'guias/crear_guia.html', {'form': form, 'tabla_productos_data_json': tabla_productos_data_json})
                
                guia.save() 

                for item_db in items_para_crear_en_db:
                    item_db.guia_envio = guia
                GuiaEnvioProducto.objects.bulk_create(items_para_crear_en_db)

                messages.success(request, 'Guía creada exitosamente!')
                return redirect('ver_guia', id=guia.id)

            except Producto.DoesNotExist:
                messages.error(request, 'Error: Uno de los productos seleccionados no existe.')
                return render(request, 'guias/crear_guia.html', {'form': form, 'error_productos': True, 'tabla_productos_data_json': tabla_productos_data_json})
            except Exception as e:
                messages.error(request, f'Error inesperado al crear la guía: {str(e)}')
                return render(request, 'guias/crear_guia.html', {'form': form, 'tabla_productos_data_json': tabla_productos_data_json})
        else:
            error_list = []
            for field, errors in form.errors.items():
                field_label = form.fields[field].label if field in form.fields and hasattr(form.fields[field], 'label') and form.fields[field].label else field
                for error in errors:
                    error_list.append(f"Error en '{field_label}': {error}")
            messages.error(request, f"Por favor corrige los errores: {'; '.join(error_list)}")
            return render(request, 'guias/crear_guia.html', {'form': form, 'tabla_productos_data_json': tabla_productos_data_json})
    else:
        form = GuiaEnvioForm()
    return render(request, 'guias/crear_guia.html', {'form': form, 'tabla_productos_data_json': tabla_productos_data_json})

@transaction.atomic
def editar_guia(request, id):
    guia = get_object_or_404(
        GuiaEnvio.objects.prefetch_related('items_guia__producto'),
        id=id
    )
    items_actuales_json = json.dumps([
        {
            'id': str(item.producto.id),
            'nombre': item.producto.nombre,
            'imagen': item.producto.imagen.url if item.producto.imagen else '',
            'precio': str(item.precio_unitario_en_guia),
            'cantidad': item.cantidad
        } for item in guia.items_guia.all()
    ])

    items_para_js = items_actuales_json

    if request.method == 'POST':
        form = GuiaEnvioForm(request.POST, instance=guia)
        
        productos_en_post_para_repopular = []
        i = 0
        while f'productos[{i}][id]' in request.POST:
            try:
                prod_id_str = request.POST[f'productos[{i}][id]']
                imagen_url = ''
                try: 
                    prod_obj_temp = Producto.objects.get(id=prod_id_str)
                    if prod_obj_temp.imagen: imagen_url = prod_obj_temp.imagen.url
                except Producto.DoesNotExist: pass

                productos_en_post_para_repopular.append({
                    'id': prod_id_str,
                    'nombre': request.POST.get(f'productos[{i}][nombre]', f'Producto ID {prod_id_str}'),
                    'precio': request.POST.get(f'productos[{i}][precio]', '0.00'),
                    'cantidad': request.POST[f'productos[{i}][cantidad]'],
                    'imagen': imagen_url
                })
            except Exception as e:
                pass
            i += 1
        
        if productos_en_post_para_repopular:
            items_para_js = json.dumps(productos_en_post_para_repopular)

        if form.is_valid():
            try:
                guia_actualizada = form.save(commit=False) 
                items_nuevos_para_db = []
                idx = 0
                hay_productos_en_tabla = False
                primer_producto_obj_de_tabla_edit = None 
                primera_cantidad_de_tabla_edit = 1      

                while f'productos[{idx}][id]' in request.POST:
                    hay_productos_en_tabla = True
                    producto_id_str = request.POST[f'productos[{idx}][id]']
                    cantidad_str = request.POST[f'productos[{idx}][cantidad]']
                    precio_unitario_str = request.POST.get(f'productos[{idx}][precio]', '0.00')

                    try:
                        producto_id = int(producto_id_str)
                        cantidad_item = int(cantidad_str)
                        precio_unitario_item = Decimal(precio_unitario_str)
                        if cantidad_item <= 0: raise ValueError("La cantidad debe ser positiva.")
                        if precio_unitario_item < 0: raise ValueError("El precio no puede ser negativo.")
                    except ValueError as ve:
                        messages.error(request, f"Datos inválidos para el producto #{idx+1} en la tabla: {ve}")
                        return render(request, 'guias/editar_guia.html', {
                            'form': form, 'guia': guia, 'items_actuales_json': items_para_js, 'error_productos': True
                        })

                    producto_obj_item = get_object_or_404(Producto, id=producto_id)
                    
                    if idx == 0: 
                        primer_producto_obj_de_tabla_edit = producto_obj_item
                        primera_cantidad_de_tabla_edit = cantidad_item

                    items_nuevos_para_db.append(
                        GuiaEnvioProducto(
                            producto=producto_obj_item,
                            cantidad=cantidad_item,
                            precio_unitario_en_guia=precio_unitario_item
                        )
                    )
                    idx += 1

                if not hay_productos_en_tabla:
                    messages.error(request, "Debe haber al menos un producto en la tabla de la guía.")
                    return render(request, 'guias/editar_guia.html', {
                        'form': form, 'guia': guia, 'items_actuales_json': items_para_js, 'error_productos': True
                    })

                if 'producto' in form.fields and primer_producto_obj_de_tabla_edit:
                     guia_actualizada.producto = primer_producto_obj_de_tabla_edit
                if 'cantidad' in form.fields:
                     guia_actualizada.cantidad = primera_cantidad_de_tabla_edit

                guia_actualizada.save() 

                guia_actualizada.items_guia.all().delete() 
                
                for item_db_obj in items_nuevos_para_db:
                    item_db_obj.guia_envio = guia_actualizada 
                
                if items_nuevos_para_db:
                    GuiaEnvioProducto.objects.bulk_create(items_nuevos_para_db)

                messages.success(request, f'Guía #{guia.id} actualizada correctamente.')
                return redirect('ver_guia', id=guia.id)
            
            except Producto.DoesNotExist:
                messages.error(request, 'Error: Uno de los productos seleccionados en la tabla ya no existe.')
            except Exception as e:
                messages.error(request, f'Error inesperado al actualizar la guía: {str(e)}')
        else: 
            error_list = []
            for field, errors_list_form in form.errors.items():
                field_label = form.fields[field].label if field in form.fields and hasattr(form.fields[field], 'label') and form.fields[field].label else field
                for error_item in errors_list_form:
                    error_list.append(f"Error en '{field_label}': {error_item}")
            messages.error(request, f"Por favor corrige los errores en el formulario: {'; '.join(error_list)}")
        
        return render(request, 'guias/editar_guia.html', {
            'form': form, 
            'guia': guia, 
            'items_actuales_json': items_para_js 
        })

    else: 
        form = GuiaEnvioForm(instance=guia)

    return render(request, 'guias/editar_guia.html', {
        'form': form,
        'guia': guia,
        'items_actuales_json': items_para_js
    })

def ver_guia(request, id):
    guia = get_object_or_404(
        GuiaEnvio.objects.select_related('producto').prefetch_related('items_guia__producto'),
        id=id
    )
    return render(request, 'guias/ver_guia.html', {
        'guia': guia,
        'estados': GuiaEnvio.ESTADO_CHOICES
    })

def eliminar_guia(request, id):
    guia = get_object_or_404(GuiaEnvio, id=id)
    if request.method == 'POST':
        guia_id_display = guia.id 
        guia.delete()
        messages.success(request, f'Guía #{guia_id_display} eliminada correctamente.')
        return redirect('listar_guias')
    return render(request, 'guias/eliminar_guia.html', {'guia': guia})

def listar_guias(request): # THIS IS THE ONLY listar_guias FUNCTION
    guias_qs = GuiaEnvio.objects.select_related(
        'producto'
    ).prefetch_related(
        'items_guia__producto'
    ).all().order_by('-fecha_creacion')

    estado_filter = request.GET.get('estado', '')
    ciudad_filter = request.GET.get('ciudad', '')
    search_query = request.GET.get('search', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')
    export_format = request.GET.get('export_format', '')

    if estado_filter:
        guias_qs = guias_qs.filter(estado=estado_filter)
    if ciudad_filter:
        guias_qs = guias_qs.filter(cliente_ciudad=ciudad_filter)

    if fecha_inicio_str:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            guias_qs = guias_qs.filter(fecha_creacion__date__gte=fecha_inicio_obj)
        except ValueError:
            messages.warning(request, "Formato de fecha de inicio inválido. Use YYYY-MM-DD.")
    if fecha_fin_str:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            guias_qs = guias_qs.filter(fecha_creacion__date__lte=fecha_fin_obj)
        except ValueError:
            messages.warning(request, "Formato de fecha de fin inválido. Use YYYY-MM-DD.")

    if search_query:
        id_query = Q()
        try:
            search_id = int(search_query)
            id_query = Q(id=search_id)
        except ValueError:
            pass

        guias_qs = guias_qs.filter(
            id_query |
            Q(cliente_nombre__icontains=search_query) |
            Q(codigo_seguimiento__icontains=search_query) |
            Q(producto__nombre__icontains=search_query) |
            Q(items_guia__producto__nombre__icontains=search_query) |
            Q(contenido_resumen__icontains=search_query)
        ).distinct()

    if export_format:
        filename_timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")

        if export_format == 'excel':
            try:
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="guias_export_{filename_timestamp}.xlsx"'
                
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Guías de Envío"

                headers = [
                    "ID", "Código Seguimiento", "Cliente", "Teléfono", "Ciudad", "Dirección", "Dirección 2",
                    "Contenido Resumen", "Observaciones", "Estado", "Total Guía",
                    "Fecha Creación", "Fecha Actualización"
                ]
                for col_num, header_title in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_num, value=header_title)
                    cell.font = Font(bold=True)
                    sheet.column_dimensions[get_column_letter(col_num)].width = 25

                for row_num, guia in enumerate(guias_qs, 2):
                    sheet.cell(row=row_num, column=1, value=guia.id)
                    sheet.cell(row=row_num, column=2, value=guia.codigo_seguimiento)
                    sheet.cell(row=row_num, column=3, value=str(guia.cliente_nombre))
                    sheet.cell(row=row_num, column=4, value=str(guia.cliente_telefono))
                    sheet.cell(row=row_num, column=5, value=str(guia.cliente_ciudad))
                    sheet.cell(row=row_num, column=6, value=str(guia.cliente_direccion)) 
                    sheet.cell(row=row_num, column=7, value=str(guia.cliente_direccion2 or ""))
                    sheet.cell(row=row_num, column=8, value=str(guia.contenido_resumen or ""))
                    sheet.cell(row=row_num, column=9, value=str(guia.observaciones or ""))
                    sheet.cell(row=row_num, column=10, value=guia.get_estado_display())
                    
                    total_val = guia.total_guia
                    if isinstance(total_val, Decimal):
                        sheet.cell(row=row_num, column=11, value=total_val).number_format = '#,##0.00'
                    elif total_val is not None:
                        try:
                            sheet.cell(row=row_num, column=11, value=Decimal(total_val)).number_format = '#,##0.00'
                        except:
                            sheet.cell(row=row_num, column=11, value=str(total_val)) 
                    else:
                        sheet.cell(row=row_num, column=11, value="")

                    sheet.cell(row=row_num, column=12, value=guia.fecha_creacion.strftime('%Y-%m-%d %H:%M') if guia.fecha_creacion else "")
                    sheet.cell(row=row_num, column=13, value=guia.fecha_actualizacion.strftime('%Y-%m-%d %H:%M') if guia.fecha_actualizacion else "")
                
                print("DEBUG: Attempting to save Excel workbook to response...")
                workbook.save(response)
                print("DEBUG: Excel workbook saved to response.")
                return response
            except Exception as e:
                print(f"!!!!!!!!!!!! EXCEL EXPORT ERROR: {e} !!!!!!!!!!!!")
                traceback.print_exc()
                messages.error(request, f"Error al generar Excel: {e}")
                return redirect('listar_guias')

        elif export_format == 'pdf':
            try:
                template_path = 'envios/guias/export_pdf_template.html' # App-namespaced path
                print(f"DEBUG: Attempting to load PDF template: {template_path}")
                template = get_template(template_path)
                
                context = {
                    'guias': guias_qs,
                    'export_date': timezone.now()
                }
                print("DEBUG: Rendering PDF HTML content...")
                html = template.render(context)

                result_file = BytesIO()
                print("DEBUG: Attempting to create PDF with pisa...")
                pisa_status = pisa.CreatePDF(html.encode('UTF-8'), dest=result_file, encoding='UTF-8')

                if pisa_status.err:
                    error_description = "Unknown error"
                    if hasattr(pisa, 'pisaErrorCodes') and pisa_status.err_code in pisa.pisaErrorCodes:
                        error_description = pisa.pisaErrorCodes[pisa_status.err_code]
                    print(f"!!!!!!!!!!!! PISA PDF Error (pisa_status.err TRUE): Code {pisa_status.err_code} - {error_description} !!!!!!!!!!!!")
                    messages.error(request, f"Error al generar PDF (código {pisa_status.err_code}). {error_description}")
                    return redirect('listar_guias')
                
                print("DEBUG: PDF creation successful (pisa_status.err is False).")
                response = HttpResponse(result_file.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="guias_export_{filename_timestamp}.pdf"'
                return response
            except Exception as e:
                print(f"!!!!!!!!!!!! PDF EXPORT GENERIC ERROR: {e} !!!!!!!!!!!!")
                traceback.print_exc()
                messages.error(request, f"Error inesperado al generar PDF: {e}")
                return redirect('listar_guias')
        
        elif export_format == 'txt':
            try:
                response = HttpResponse(content_type='text/plain; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="guias_export_{filename_timestamp}.txt"'
                
                writer = csv.writer(response, delimiter='\t')
                headers_txt = [
                    "ID", "Código Seguimiento", "Cliente", "Teléfono", "Ciudad", "Dirección",
                    "Contenido Resumen", "Estado", "Total Guía", "Fecha Creación"
                ]
                writer.writerow(headers_txt)

                for guia in guias_qs:
                    writer.writerow([
                        guia.id,
                        guia.codigo_seguimiento,
                        str(guia.cliente_nombre),
                        str(guia.cliente_telefono),
                        str(guia.cliente_ciudad),
                        str(guia.cliente_direccion),
                        str(guia.contenido_resumen or ""),
                        guia.get_estado_display(),
                        str(guia.total_guia or ""),
                        guia.fecha_creacion.strftime('%Y-%m-%d %H:%M') if guia.fecha_creacion else ""
                    ])
                print("DEBUG: TXT export generated.")
                return response
            except Exception as e:
                print(f"!!!!!!!!!!!! TXT EXPORT ERROR: {e} !!!!!!!!!!!!")
                traceback.print_exc()
                messages.error(request, f"Error al generar archivo TXT: {e}")
                return redirect('listar_guias')

    # For regular HTML rendering (if not exporting)
    todas_las_ciudades_en_db = GuiaEnvio.objects.values_list('cliente_ciudad', flat=True).distinct().order_by('cliente_ciudad')
    ciudades_dropdown_choices = [(ciudad, ciudad) for ciudad in todas_las_ciudades_en_db if ciudad]

    context = {
        'guias': guias_qs,
        'estados': GuiaEnvio.ESTADO_CHOICES,
        'ciudades': ciudades_dropdown_choices,
        'estado_selected': estado_filter,
        'ciudad_selected': ciudad_filter,
        'search_query': search_query,
        'fecha_inicio_selected': fecha_inicio_str,
        'fecha_fin_selected': fecha_fin_str,
    }
    return render(request, 'guias/listar_guias.html', context)
    guias_qs = GuiaEnvio.objects.select_related(
        'producto'
    ).prefetch_related(
        'items_guia__producto'
    ).all().order_by('-fecha_creacion')

    estado_filter = request.GET.get('estado', '')
    ciudad_filter = request.GET.get('ciudad', '')
    search_query = request.GET.get('search', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')
    export_format = request.GET.get('export_format', '')

    if estado_filter:
        guias_qs = guias_qs.filter(estado=estado_filter)
    if ciudad_filter:
        guias_qs = guias_qs.filter(cliente_ciudad=ciudad_filter)

    if fecha_inicio_str:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            guias_qs = guias_qs.filter(fecha_creacion__date__gte=fecha_inicio_obj)
        except ValueError:
            messages.warning(request, "Formato de fecha de inicio inválido. Use YYYY-MM-DD.")
    if fecha_fin_str:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            guias_qs = guias_qs.filter(fecha_creacion__date__lte=fecha_fin_obj)
        except ValueError:
            messages.warning(request, "Formato de fecha de fin inválido. Use YYYY-MM-DD.")

    if search_query:
        id_query = Q()
        try:
            search_id = int(search_query)
            id_query = Q(id=search_id)
        except ValueError:
            pass

        guias_qs = guias_qs.filter(
            id_query |
            Q(cliente_nombre__icontains=search_query) |
            Q(codigo_seguimiento__icontains=search_query) |
            Q(producto__nombre__icontains=search_query) |
            Q(items_guia__producto__nombre__icontains=search_query) |
            Q(contenido_resumen__icontains=search_query)
        ).distinct()

    if export_format:
        filename_timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")

        if export_format == 'excel':
            try:
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="guias_export_{filename_timestamp}.xlsx"'
                
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Guías de Envío"

                headers = [
                    "ID", "Código Seguimiento", "Cliente", "Teléfono", "Ciudad", "Dirección", "Dirección 2",
                    "Contenido Resumen", "Observaciones", "Estado", "Total Guía",
                    "Fecha Creación", "Fecha Actualización"
                ]
                for col_num, header_title in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_num, value=header_title)
                    cell.font = Font(bold=True)
                    sheet.column_dimensions[get_column_letter(col_num)].width = 25

                for row_num, guia in enumerate(guias_qs, 2):
                    sheet.cell(row=row_num, column=1, value=guia.id)
                    sheet.cell(row=row_num, column=2, value=guia.codigo_seguimiento)
                    sheet.cell(row=row_num, column=3, value=str(guia.cliente_nombre))
                    sheet.cell(row=row_num, column=4, value=str(guia.cliente_telefono))
                    sheet.cell(row=row_num, column=5, value=str(guia.cliente_ciudad))
                    sheet.cell(row=row_num, column=6, value=str(guia.cliente_direccion)) 
                    sheet.cell(row=row_num, column=7, value=str(guia.cliente_direccion2 or ""))
                    sheet.cell(row=row_num, column=8, value=str(guia.contenido_resumen or ""))
                    sheet.cell(row=row_num, column=9, value=str(guia.observaciones or ""))
                    sheet.cell(row=row_num, column=10, value=guia.get_estado_display())
                    
                    total_val = guia.total_guia
                    if isinstance(total_val, Decimal):
                        sheet.cell(row=row_num, column=11, value=total_val).number_format = '#,##0.00'
                    elif total_val is not None:
                        try:
                            sheet.cell(row=row_num, column=11, value=Decimal(total_val)).number_format = '#,##0.00'
                        except:
                            sheet.cell(row=row_num, column=11, value=str(total_val)) 
                    else:
                        sheet.cell(row=row_num, column=11, value="")

                    sheet.cell(row=row_num, column=12, value=guia.fecha_creacion.strftime('%Y-%m-%d %H:%M') if guia.fecha_creacion else "")
                    sheet.cell(row=row_num, column=13, value=guia.fecha_actualizacion.strftime('%Y-%m-%d %H:%M') if guia.fecha_actualizacion else "")
                
                print("DEBUG: Attempting to save Excel workbook to response...")
                workbook.save(response)
                print("DEBUG: Excel workbook saved to response.")
                return response
            except Exception as e:
                print(f"!!!!!!!!!!!! EXCEL EXPORT ERROR: {e} !!!!!!!!!!!!")
                traceback.print_exc()
                messages.error(request, f"Error al generar Excel: {e}")
                return redirect('listar_guias')

        elif export_format == 'pdf':
            try:
                template_path = 'envios/guias/export_pdf_template.html'
                print(f"DEBUG: Attempting to load PDF template: {template_path}")
                template = get_template(template_path)
                
                context = {
                    'guias': guias_qs,
                    'export_date': timezone.now()
                }
                print("DEBUG: Rendering PDF HTML content...")
                html = template.render(context)

                result_file = BytesIO()
                print("DEBUG: Attempting to create PDF with pisa...")
                pisa_status = pisa.CreatePDF(html.encode('UTF-8'), dest=result_file, encoding='UTF-8')

                if pisa_status.err:
                    print(f"!!!!!!!!!!!! PISA PDF Error (pisa_status.err TRUE): {pisa_status.err_code} - {pisa.pisaErrorCodes[pisa_status.err_code]} !!!!!!!!!!!!")
                    messages.error(request, f"Error al generar PDF (código {pisa_status.err_code}).")
                    return redirect('listar_guias')
                
                print("DEBUG: PDF creation successful (pisa_status.err is False).")
                response = HttpResponse(result_file.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="guias_export_{filename_timestamp}.pdf"'
                return response
            except Exception as e:
                print(f"!!!!!!!!!!!! PDF EXPORT GENERIC ERROR: {e} !!!!!!!!!!!!")
                traceback.print_exc()
                messages.error(request, f"Error inesperado al generar PDF: {e}")
                return redirect('listar_guias')
        
        elif export_format == 'txt':
            try:
                response = HttpResponse(content_type='text/plain; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="guias_export_{filename_timestamp}.txt"'
                
                writer = csv.writer(response, delimiter='\t')
                headers_txt = [
                    "ID", "Código Seguimiento", "Cliente", "Teléfono", "Ciudad", "Dirección",
                    "Contenido Resumen", "Estado", "Total Guía", "Fecha Creación"
                ]
                writer.writerow(headers_txt)

                for guia in guias_qs:
                    writer.writerow([
                        guia.id,
                        guia.codigo_seguimiento,
                        str(guia.cliente_nombre),
                        str(guia.cliente_telefono),
                        str(guia.cliente_ciudad),
                        str(guia.cliente_direccion),
                        str(guia.contenido_resumen or ""),
                        guia.get_estado_display(),
                        str(guia.total_guia or ""),
                        guia.fecha_creacion.strftime('%Y-%m-%d %H:%M') if guia.fecha_creacion else ""
                    ])
                print("DEBUG: TXT export generated.")
                return response
            except Exception as e:
                print(f"!!!!!!!!!!!! TXT EXPORT ERROR: {e} !!!!!!!!!!!!")
                traceback.print_exc()
                messages.error(request, f"Error al generar archivo TXT: {e}")
                return redirect('listar_guias')

    todas_las_ciudades_en_db = GuiaEnvio.objects.values_list('cliente_ciudad', flat=True).distinct().order_by('cliente_ciudad')
    ciudades_dropdown_choices = [(ciudad, ciudad) for ciudad in todas_las_ciudades_en_db if ciudad]

    context = {
        'guias': guias_qs,
        'estados': GuiaEnvio.ESTADO_CHOICES,
        'ciudades': ciudades_dropdown_choices,
        'estado_selected': estado_filter,
        'ciudad_selected': ciudad_filter,
        'search_query': search_query,
        'fecha_inicio_selected': fecha_inicio_str,
        'fecha_fin_selected': fecha_fin_str,
    }
    return render(request, 'guias/listar_guias.html', context)

    


# --- Campaign Views ---
def index(request):
    context = {
        'countries': [
            {'code': 'COL', 'name': 'Colombia'},
            {'code': 'ECU', 'name': 'Ecuador'},
            {'code': 'VEN', 'name': 'Venezuela'}
        ],
        'platforms': [
            {'name': 'Facebook', 'code': 'F'},
            {'name': 'TikTok', 'code': 'T'}
        ],
        'campaign_types': [
            {'name': 'CBO', 'desc': 'Campaign Budget Optimization'},
            {'name': 'ABO', 'desc': 'Ad Set Budget'},
            {'name': 'FLEXIBLE', 'desc': 'Flexible'}
        ],
        'persons': [
            {'initials': 'DP', 'name': 'Duvan Pantoja'},
            {'initials': 'CP', 'name': 'Carolina Pantoja'},
            {'initials': 'KO', 'name': 'Karen Orbes'}
            {'initials': 'EM', 'name': 'Emili Pantoja'}
        ]
    }
    return render(request, 'campaigns/index.html', context)

    return render(request, 'campaigns/index.html', context)

def saved_campaigns(request):
    # Aquí ya no se usa localStorage, se leen de la BD
    campaigns_from_db = Campaign.objects.all().order_by('-created_at')
    
    # Si aún quieres pasar los datos parseados al JS para que `script_campanas.js`
    # los maneje como antes (aunque el filtrado ahora debería ser server-side idealmente)
    # o si `script_campanas.js` solo hace display.
    # Pasaremos los gui_strings para que el JS los parsee.
    
    # Convertir las campañas de la base de datos a una lista de diccionarios o solo los gui_strings
    # para que el JavaScript `script_campanas.js` pueda procesarlos.
    # Esto es si quieres mantener la lógica de parseo y renderizado en el JS.
    # Alternativamente, podrías pasar los objetos Campaign completos y renderizar en la plantilla Django.
    
    # Opción 1: Pasar solo los gui_strings (como parece que espera tu script_campanas.js)
    # django_campanas_data = [c.gui_string for c in campaigns_from_db]
    
    # Opción 2: Pasar objetos Campaign completos si quieres más datos en la plantilla
    context = {
        'campaigns': campaigns_from_db, # Para iterar en la plantilla Django directamente
        # 'django_campanas_json': json.dumps(django_campanas_data) # Para que JS lo tome
    }
    return render(request, 'campaigns/campanas_guardadas.html', context)


@csrf_exempt
def save_campaign_api(request): # Renombrada para evitar conflicto
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            gui_string = data.get('gui_string')
            if not gui_string:
                return JsonResponse({'success': False, 'error': 'gui_string es requerido'})

            # Aquí puedes añadir más validaciones para los otros campos si es necesario
            # antes de crear el objeto Campaign.

            # Si ya existe, podrías actualizarla o devolver un error, según tu lógica.
            # Por ahora, si existe, devolvemos error (como en tu código original).
            if Campaign.objects.filter(gui_string=gui_string).exists():
                return JsonResponse({'success': False, 'error': 'Esta campaña (GUI string) ya existe'})
            
            campaign = Campaign(
                gui_string=gui_string,
                pais=data.get('pais'),
                plataforma=data.get('plataforma'),
                tipo_campana=data.get('tipo_campana'),
                fecha=data.get('fecha'), 
                nombre_reloj=data.get('nombre_reloj'),
                id_reloj=data.get('id_reloj'),
                persona=data.get('persona'),
                num_importacion=data.get('num_importacion')
            )
            campaign.save()
            return JsonResponse({'success': True, 'message': 'Campaña guardada exitosamente en la base de datos'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'JSON inválido'})
        except Exception as e:
            # Es bueno loggear el error en el servidor para depuración
            print(f"Error al guardar campaña: {e}")
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': f'Error interno del servidor: {str(e)}'})
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

@csrf_exempt
def clear_campaigns_api(request): # Renombrada
    if request.method == 'POST':
        try:
            count, _ = Campaign.objects.all().delete()
            return JsonResponse({'success': True, 'message': f'{count} campañas han sido eliminadas de la base de datos'})
        except Exception as e:
            print(f"Error al limpiar campañas: {e}")
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': f'Error interno del servidor: {str(e)}'})
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)